"""
Modul untuk menangani penyimpanan data ke SQLite dan export ke Excel
"""

import sqlite3
import logging
import os
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from models.data_models import ScrapedData


class DatabaseHandler:
    """Kelas untuk menangani penyimpanan dan pengambilan data dari SQLite"""

    DB_PATH = os.path.join("results", "scraped_data.db")

    @staticmethod
    def _init_db():
        """Inisialisasi database SQLite"""
        os.makedirs("results", exist_ok=True)
        conn = sqlite3.connect(DatabaseHandler.DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS scraped_data (
                title TEXT,
                provider TEXT,
                size TEXT,
                status TEXT,
                download_url TEXT,
                bypass_url TEXT,
                target_code TEXT,
                PRIMARY KEY (title, provider, target_code)
            )
            """
        )
        conn.commit()
        conn.close()

    @staticmethod
    def save_to_sqlite(data: List[ScrapedData]) -> int:
        """Menyimpan data ke SQLite dan mengembalikan jumlah item baru"""
        DatabaseHandler._init_db()
        conn = sqlite3.connect(DatabaseHandler.DB_PATH)
        cursor = conn.cursor()
        new_items = 0

        for item in data:
            cursor.execute(
                """
                INSERT OR IGNORE INTO scraped_data
                (title, provider, size, status, download_url, bypass_url, target_code)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    item.title,
                    item.provider,
                    item.size,
                    item.status,
                    item.download_url,
                    item.bypass_url,
                    item.target_code,
                ),
            )
            if cursor.rowcount > 0:
                new_items += 1

        conn.commit()
        conn.close()
        if new_items > 0:
            logging.info(f"[DB] Disimpan {new_items} item ke database")
        return new_items

    @staticmethod
    def is_data_exists(title: str, provider: str, target_code: str) -> bool:
        conn = sqlite3.connect(DatabaseHandler.DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT 1 FROM scraped_data
            WHERE title = ? AND provider = ? AND target_code = ?
            """,
            (title, provider, target_code),
        )
        exists = cursor.fetchone() is not None
        conn.close()
        return exists

    @staticmethod
    def get_data_by_title_provider(
        title: str, provider: str, target_code: str
    ) -> ScrapedData | None:
        conn = sqlite3.connect(DatabaseHandler.DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT title, provider, size, status, download_url, bypass_url, target_code
            FROM scraped_data
            WHERE title = ? AND provider = ? AND target_code = ?
            """,
            (title, provider, target_code),
        )
        row = cursor.fetchone()
        conn.close()
        if row:
            return ScrapedData(
                title=row[0],
                provider=row[1],
                size=row[2],
                status=row[3],
                download_url=row[4],
                bypass_url=row[5],
                target_code=row[6],
                container_title="N/A",
            )
        return None

    @staticmethod
    def get_all_data() -> List[ScrapedData]:
        conn = sqlite3.connect(DatabaseHandler.DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT title, provider, size, status, download_url, bypass_url, target_code
            FROM scraped_data
            """
        )
        rows = cursor.fetchall()
        conn.close()
        return [
            ScrapedData(
                title=row[0],
                provider=row[1],
                size=row[2],
                status=row[3],
                download_url=row[4],
                bypass_url=row[5],
                target_code=row[6],
                container_title="N/A",
            )
            for row in rows
        ]

    @staticmethod
    def get_data_by_target_code(target_code: str) -> List[ScrapedData]:
        conn = sqlite3.connect(DatabaseHandler.DB_PATH)
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT title, provider, size, status, download_url, bypass_url, target_code
            FROM scraped_data
            WHERE target_code = ?
            """,
            (target_code,),
        )
        rows = cursor.fetchall()
        conn.close()
        return [
            ScrapedData(
                title=row[0],
                provider=row[1],
                size=row[2],
                status=row[3],
                download_url=row[4],
                bypass_url=row[5],
                target_code=row[6],
                container_title="N/A",
            )
            for row in rows
        ]

    @staticmethod
    def _is_sheet_empty(sheet: Worksheet) -> bool:
        return sheet.max_row <= 1

    @staticmethod
    def export_to_excel(data: List[ScrapedData], target_code: str):
        """Export data ke file Excel RESULT_DATABASE.xlsx"""
        if not data and not target_code:
            return

        os.makedirs("results", exist_ok=True)
        filename = os.path.join("results", "RESULT_DATABASE.xlsx")
        relevant_data = DatabaseHandler.get_data_by_target_code(target_code)
        combined_data = relevant_data[:]
        input_keys = {
            (
                item.title.lower().strip(),
                item.provider.lower().strip(),
                item.target_code.lower().strip(),
            )
            for item in data
        }

        for item in data:
            item_key = (
                item.title.lower().strip(),
                item.provider.lower().strip(),
                item.target_code.lower().strip(),
            )
            if item_key not in {
                (
                    d.title.lower().strip(),
                    d.provider.lower().strip(),
                    d.target_code.lower().strip(),
                )
                for d in relevant_data
            }:
                combined_data.append(item)

        if not combined_data:
            return

        try:
            if os.path.exists(filename):
                wb = load_workbook(filename)
                for sheet_name in wb.sheetnames[:]:
                    sheet = wb[sheet_name]
                    if (
                        DatabaseHandler._is_sheet_empty(sheet)
                        or sheet_name.lower() == "sheet"
                    ):
                        wb.remove(sheet)
            else:
                wb = Workbook()
                wb.remove(wb.active)

            all_sheet_name = "ALL_PROVIDER_RESULTS"
            if all_sheet_name not in wb.sheetnames:
                all_sheet = wb.create_sheet(all_sheet_name, 0)
                all_sheet.append(
                    [
                        "No",
                        "Title",
                        "Provider",
                        "Size",
                        "Status",
                        "Download URL",
                        "Bypass URL",
                        "_target_code",
                    ]
                )
            else:
                all_sheet = wb[all_sheet_name]

            existing_all_entries = set()
            for row in all_sheet.iter_rows(min_row=2, values_only=True):
                if row[1] and row[2] and row[7]:
                    existing_all_entries.add(
                        (
                            str(row[1]).lower().strip(),
                            str(row[2]).lower().strip(),
                            str(row[7]).lower().strip(),
                        )
                    )

            new_all_items = 0
            for item in combined_data:
                item_key = (
                    item.title.lower().strip(),
                    item.provider.lower().strip(),
                    item.target_code.lower().strip(),
                )
                if item_key not in existing_all_entries:
                    all_sheet.append(
                        [
                            all_sheet.max_row,
                            item.title,
                            item.provider,
                            item.size,
                            item.status,
                            item.download_url,
                            item.bypass_url,
                            item.target_code,
                        ]
                    )
                    existing_all_entries.add(item_key)
                    new_all_items += 1

            for idx, row in enumerate(all_sheet.iter_rows(min_row=2), 1):
                row[0].value = idx

            if DatabaseHandler._is_sheet_empty(all_sheet):
                wb.remove(all_sheet)

            providers = {item.provider for item in combined_data}
            for provider in providers:
                provider_items = [
                    item for item in combined_data if item.provider == provider
                ]
                provider_new_items = False
                existing_provider_entries = set()
                provider_sheet_name = provider[:31]
                if provider_sheet_name in wb.sheetnames:
                    provider_sheet = wb[provider_sheet_name]
                    for row in provider_sheet.iter_rows(min_row=2, values_only=True):
                        if row[1] and row[2] and row[7]:
                            existing_provider_entries.add(
                                (
                                    str(row[1]).lower().strip(),
                                    str(row[2]).lower().strip(),
                                    str(row[7]).lower().strip(),
                                )
                            )

                for item in provider_items:
                    item_key = (
                        item.title.lower().strip(),
                        item.provider.lower().strip(),
                        item.target_code.lower().strip(),
                    )
                    if item_key not in existing_provider_entries:
                        provider_new_items = True
                        break

                if provider_new_items:
                    if provider_sheet_name not in wb.sheetnames:
                        provider_sheet = wb.create_sheet(provider_sheet_name)
                        provider_sheet.append(
                            [
                                "No",
                                "Title",
                                "Provider",
                                "Size",
                                "Status",
                                "Download URL",
                                "Bypass URL",
                                "_target_code",
                            ]
                        )
                    else:
                        provider_sheet = wb[provider_sheet_name]

                    for item in provider_items:
                        item_key = (
                            item.title.lower().strip(),
                            item.provider.lower().strip(),
                            item.target_code.lower().strip(),
                        )
                        if item_key not in existing_provider_entries:
                            provider_sheet.append(
                                [
                                    provider_sheet.max_row,
                                    item.title,
                                    item.provider,
                                    item.size,
                                    item.status,
                                    item.download_url,
                                    item.bypass_url,
                                    item.target_code,
                                ]
                            )
                            existing_provider_entries.add(item_key)

                    for idx, row in enumerate(provider_sheet.iter_rows(min_row=2), 1):
                        row[0].value = idx

                    if DatabaseHandler._is_sheet_empty(provider_sheet):
                        wb.remove(provider_sheet)

            has_data = False
            for sheet_name in wb.sheetnames[:]:
                sheet = wb[sheet_name]
                if not DatabaseHandler._is_sheet_empty(sheet):
                    has_data = True
                else:
                    wb.remove(sheet)

            if not has_data or not wb.sheetnames:
                return

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                column_widths = {
                    "A": 5,
                    "B": 60,
                    "C": 15,
                    "D": 10,
                    "E": 15,
                    "F": 60,
                    "G": 60,
                    "H": 15,
                }
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

            wb.save(filename)

        except Exception as e:
            logging.error(f"[ERROR] Gagal menyimpan Excel: {str(e)}")
