"""
Modul untuk menangani penyimpanan file individual
"""

import os
import re
import logging
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from models.data_models import ScrapedData
from core.database import DatabaseHandler


class FileHandler:
    """Kelas untuk menangani penyimpanan file individual"""

    @staticmethod
    def _sanitize_filename(title: str, target_code: str = None) -> str:
        """
        Membersihkan nama file dan menghindari karakter tidak valid.
        Mengembalikan target_code hanya jika title benar-benar tidak valid.
        """
        logging.debug(f"📁⠀ Sanitizing title: {title}, target_code: {target_code}")
        if not title or title.strip().lower() in ("n/a", "unknown", ""):
            logging.warning(
                f"📁⠀ Title tidak valid ({title}), menggunakan target_code: {target_code}"
            )
            return target_code if target_code else "Untitled"

        # Ganti karakter tidak valid, kecuali titik
        cleaned_title = re.sub(r'[\/\\:*?"<>|]', "", title.strip())
        # Ganti beberapa spasi dengan titik
        cleaned_title = re.sub(r"\s+", ".", cleaned_title)
        max_length = 200
        cleaned_title = cleaned_title[:max_length]

        if not cleaned_title:
            logging.warning(
                f"📁⠀ Title kosong setelah pembersihan, menggunakan target_code: {target_code}"
            )
            return target_code if target_code else "Untitled"

        logging.debug(f"📁⠀ Sanitized title: {cleaned_title}")
        return cleaned_title

    @staticmethod
    def _is_sheet_empty(sheet: Worksheet) -> bool:
        return sheet.max_row <= 1

    @staticmethod
    def save_individual_files(
        data: List[ScrapedData], target_code: str, container_title: str = None
    ):
        """Menyimpan file individual berdasarkan container_title"""
        if not data or not target_code:
            logging.warning("📁⠀ Tidak ada data atau target_code untuk disimpan")
            return

        os.makedirs("results", exist_ok=True)
        relevant_data = DatabaseHandler.get_data_by_target_code(target_code)
        combined_data = relevant_data[:]
        input_keys = {
            (
                item.title.lower().strip(),
                item.provider.lower().strip(),
                item.target_code.lower().strip(),
            )
            for item in data
        }

        # Gunakan container_title dari parameter jika ada, jika tidak cari dari data
        selected_title = container_title
        if not selected_title:
            for item in data:
                if (
                    item.container_title
                    and item.container_title.strip().lower()
                    not in ("n/a", "unknown", "")
                ):
                    selected_title = item.container_title
                    break
            if not selected_title:
                for item in combined_data:
                    if (
                        item.container_title
                        and item.container_title.strip().lower()
                        not in ("n/a", "unknown", "")
                    ):
                        selected_title = item.container_title
                        break
            if not selected_title:
                logging.warning(
                    f"📁⠀ Tidak ada container_title valid untuk target_code {target_code}, menggunakan target_code"
                )
                selected_title = target_code

        # Tambahkan data baru ke combined_data
        for item in data:
            item_key = (
                item.title.lower().strip(),
                item.provider.lower().strip(),
                item.target_code.lower().strip(),
            )
            if item_key not in {
                (
                    d.title.lower().strip(),
                    d.provider.lower().strip(),
                    d.target_code.lower().strip(),
                )
                for d in relevant_data
            }:
                item.container_title = selected_title
                combined_data.append(item)

        if not combined_data:
            logging.warning(
                f"📁⠀ Tidak ada data untuk disimpan untuk target_code: {target_code}"
            )
            return

        filename = os.path.join(
            "results",
            f"{FileHandler._sanitize_filename(selected_title, target_code)}.xlsx",
        )
        logging.info(f"📁 Menyimpan file ke: {filename}")

        try:
            if os.path.exists(filename):
                wb = load_workbook(filename)
                for sheet_name in wb.sheetnames[:]:
                    sheet = wb[sheet_name]
                    if (
                        FileHandler._is_sheet_empty(sheet)
                        or sheet_name.lower() == "sheet"
                    ):
                        wb.remove(sheet)
            else:
                wb = Workbook()
                wb.remove(wb.active)

            all_sheet_name = "ALL_PROVIDER_RESULTS"
            if all_sheet_name not in wb.sheetnames:
                all_sheet = wb.create_sheet(all_sheet_name, 0)
                all_sheet.append(
                    [
                        "No",
                        "Title",
                        "Provider",
                        "Size",
                        "Status",
                        "Download URL",
                        "Bypass URL",
                        "_target_code",
                    ]
                )
            else:
                all_sheet = wb[all_sheet_name]

            existing_all_entries = set()
            for row in all_sheet.iter_rows(min_row=2, values_only=True):
                if row[1] and row[2] and row[7]:
                    existing_all_entries.add(
                        (
                            str(row[1]).lower().strip(),
                            str(row[2]).lower().strip(),
                            str(row[7]).lower().strip(),
                        )
                    )

            new_all_items = 0
            for item in combined_data:
                item_key = (
                    item.title.lower().strip(),
                    item.provider.lower().strip(),
                    item.target_code.lower().strip(),
                )
                if item_key not in existing_all_entries:
                    all_sheet.append(
                        [
                            all_sheet.max_row,
                            item.title,
                            item.provider,
                            item.size,
                            item.status,
                            item.download_url,
                            item.bypass_url,
                            item.target_code,
                        ]
                    )
                    existing_all_entries.add(item_key)
                    new_all_items += 1

            for idx, row in enumerate(all_sheet.iter_rows(min_row=2), 1):
                row[0].value = idx

            if FileHandler._is_sheet_empty(all_sheet):
                wb.remove(all_sheet)

            providers = {item.provider for item in combined_data}
            for provider in providers:
                provider_items = [
                    item for item in combined_data if item.provider == provider
                ]
                provider_new_items = False
                existing_provider_entries = set()
                provider_sheet_name = provider[:31]
                if provider_sheet_name in wb.sheetnames:
                    provider_sheet = wb[provider_sheet_name]
                    for row in provider_sheet.iter_rows(min_row=2, values_only=True):
                        if row[1] and row[2] and row[7]:
                            existing_provider_entries.add(
                                (
                                    str(row[1]).lower().strip(),
                                    str(row[2]).lower().strip(),
                                    str(row[7]).lower().strip(),
                                )
                            )

                for item in provider_items:
                    item_key = (
                        item.title.lower().strip(),
                        item.provider.lower().strip(),
                        item.target_code.lower().strip(),
                    )
                    if item_key not in existing_provider_entries:
                        provider_new_items = True
                        break

                if provider_new_items:
                    if provider_sheet_name not in wb.sheetnames:
                        provider_sheet = wb.create_sheet(provider_sheet_name)
                        provider_sheet.append(
                            [
                                "No",
                                "Title",
                                "Provider",
                                "Size",
                                "Status",
                                "Download URL",
                                "Bypass URL",
                                "_target_code",
                            ]
                        )
                    else:
                        provider_sheet = wb[provider_sheet_name]

                    for item in provider_items:
                        item_key = (
                            item.title.lower().strip(),
                            item.provider.lower().strip(),
                            item.target_code.lower().strip(),
                        )
                        if item_key not in existing_provider_entries:
                            provider_sheet.append(
                                [
                                    provider_sheet.max_row,
                                    item.title,
                                    item.provider,
                                    item.size,
                                    item.status,
                                    item.download_url,
                                    item.bypass_url,
                                    item.target_code,
                                ]
                            )
                            existing_provider_entries.add(item_key)

                    for idx, row in enumerate(provider_sheet.iter_rows(min_row=2), 1):
                        row[0].value = idx

                    if FileHandler._is_sheet_empty(provider_sheet):
                        wb.remove(provider_sheet)

            has_data = False
            for sheet_name in wb.sheetnames[:]:
                sheet = wb[sheet_name]
                if not FileHandler._is_sheet_empty(sheet):
                    has_data = True
                else:
                    wb.remove(sheet)

            if not has_data or not wb.sheetnames:
                logging.warning(f"📁⠀ Tidak ada data untuk disimpan ke {filename}")
                return

            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            for sheet in wb.sheetnames:
                ws = wb[sheet]
                column_widths = {
                    "A": 5,
                    "B": 60,
                    "C": 15,
                    "D": 10,
                    "E": 15,
                    "F": 60,
                    "G": 60,
                    "H": 15,
                }
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

            wb.save(filename)

        except Exception as e:
            logging.error(
                f"[ERROR] Gagal menyimpan file individual ke {filename}: {str(e)}"
            )
