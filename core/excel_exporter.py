import os
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from typing import List, Dict, Any


class ExcelExporter:
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]]):
        """Export data from SQLite to Excel format"""
        if not data:
            logging.info("[EXCEL] Tidak ada data untuk diexport")
            return

        os.makedirs("results", exist_ok=True)
        db_file = "results/RESULT_DATABASE.xlsx"

        try:
            # 1. Inisialisasi Workbook
            if os.path.exists(db_file):
                wb = load_workbook(db_file)
                if "ALL_PROVIDER_RESULTS" not in wb.sheetnames:
                    wb.create_sheet("ALL_PROVIDER_RESULTS", 0)
                    ExcelExporter._init_sheet(wb["ALL_PROVIDER_RESULTS"])
            else:
                wb = Workbook()
                wb.remove(wb.active)
                wb.create_sheet("ALL_PROVIDER_RESULTS", 0)
                ExcelExporter._init_sheet(wb["ALL_PROVIDER_RESULTS"])

            all_sheet = wb["ALL_PROVIDER_RESULTS"]

            # 2. Kelompokkan data berdasarkan target_code
            target_groups = {}
            for item in data:
                target_groups.setdefault(item["target_code"], []).append(item)

            # 3. Simpan data ke Excel
            for target_code, items in target_groups.items():
                # Tambahkan separator jika target_code berbeda
                if (
                    all_sheet.max_row > 1
                    and str(all_sheet.cell(row=all_sheet.max_row, column=8).value)
                    != target_code
                ):
                    all_sheet.append([None] * 8)

                # Simpan ke sheet utama
                for idx, item in enumerate(items, start=all_sheet.max_row):
                    all_sheet.append(
                        [
                            idx,
                            item["title"],
                            item["provider"],
                            item["size"],
                            item["status"],
                            item["download_url"],
                            item["bypass_url"],
                            target_code,
                        ]
                    )

                # Simpan ke sheet provider
                for item in items:
                    provider_sheet_name = item["provider"][:31]
                    if provider_sheet_name not in wb.sheetnames:
                        wb.create_sheet(provider_sheet_name)
                        ExcelExporter._init_sheet(wb[provider_sheet_name])

                    provider_sheet = wb[provider_sheet_name]
                    provider_sheet.append(
                        [
                            provider_sheet.max_row,
                            item["title"],
                            item["provider"],
                            item["size"],
                            item["status"],
                            item["download_url"],
                            item["bypass_url"],
                            target_code,
                        ]
                    )

            # 4. Formatting
            ExcelExporter._format_workbook(wb)
            wb.save(db_file)
            logging.info(f"[EXCEL] Data berhasil disimpan ke {db_file}")

        except Exception as e:
            logging.error(f"[EXCEL] Gagal menyimpan data: {str(e)}")
            raise

    @staticmethod
    def _init_sheet(sheet):
        """Initialize sheet with headers"""
        sheet.append(
            [
                "No",
                "Title",
                "Provider",
                "Size",
                "Status",
                "Download URL",
                "Bypass URL",
                "_target_code",
            ]
        )

    @staticmethod
    def _format_workbook(wb):
        """Apply formatting to all sheets"""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            column_widths = {
                "A": 5,
                "B": 60,
                "C": 15,
                "D": 10,
                "E": 15,
                "F": 60,
                "G": 60,
                "H": 15,
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
