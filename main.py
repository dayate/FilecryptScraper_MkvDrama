"""
Program utama FileCrypt Scraper
"""

import os
import sys
import logging
import time
from tabulate import tabulate
import tkinter as tk
from tkinter import filedialog
import sqlite3
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from typing import Dict, List, Optional
from rich.console import Console
from rich.table import Table
from rich.text import Text
from rich.panel import Panel
from core.browser import BrowserManager
from core.scraper import FileCryptScraper
from core.database import DatabaseHandler
from core.file_handler import FileHandler
from core.logger import setup_logging
from config.settings import DEFAULT_CONFIG

# Inisialisasi rich console
console = Console()


def select_input_method() -> str:
    """Meminta pengguna memilih metode input URL atau cek database"""
    content = (
        "1. Masukkan satu URL langsung\n"
        "2. Pilih file berisi daftar URL (.txt)\n"
        "3. Cek database"
    )
    console.print(
        Panel(
            content,
            title="[bold blue]PILIH METODE[/bold blue]",
            border_style="cyan",
            expand=True,
        )
    )

    while True:
        try:
            choice = console.input("Masukkan nomor opsi (1-3): ").strip()
            if choice in ["1", "2", "3"]:
                return choice
            console.print(
                "⚠️ [yellow]Opsi tidak valid! Harap masukkan nomor 1, 2, atau 3.[/yellow]"
            )
        except Exception as e:
            logging.error(f"[ERROR] Error memilih metode input: {str(e)}")
            console.print(
                f"❌ [red]Terjadi error! Harap masukkan nomor 1, 2, atau 3.[/red]"
            )
        time.sleep(1)


def get_valid_url() -> str:
    """Meminta satu URL valid dari pengguna berdasarkan pola yang ada di config."""
    valid_patterns = DEFAULT_CONFIG.scraper.valid_urls
    if not valid_patterns:
        console.print(
            "[bold red]⚠️ Error: Tidak ada URL valid yang dikonfigurasi di 'config.yaml'.[/bold red]"
        )
        return ""

    while True:
        try:
            url = console.input("Masukkan URL: ").strip()
            # Memeriksa apakah URL cocok dengan salah satu pola valid dan diakhiri dengan .html
            if any(
                url.startswith(pattern) for pattern in valid_patterns
            ) and url.endswith(".html"):
                return url
            console.print(
                "⚠️ [yellow]URL tidak valid! Harap masukkan URL yang benar.[/yellow]"
            )
        except (KeyboardInterrupt, EOFError):
            console.print(
                "\n [bold yellow]Proses dibatalkan oleh pengguna.[/bold yellow]"
            )
            return ""
        except Exception as e:
            logging.error(f"❌ Error saat memasukkan URL: {str(e)}")
            console.print(
                f"❌ [red]Terjadi error! Harap masukkan URL yang valid.[/red]"
            )
        time.sleep(1)


def get_urls_from_file() -> List[str]:
    """Membuka file explorer dan membaca URL dari file .txt"""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Pilih file .txt berisi URL", filetypes=[("Text files", "*.txt")]
    )
    root.destroy()

    if not file_path:
        logging.error("❌⠀ Tidak ada file yang dipilih")
        console.print("❌⠀ [red]Tidak ada file yang dipilih. Program berhenti.[/red]")
        sys.exit(1)

    valid_urls = []
    valid_patterns = DEFAULT_CONFIG.scraper.valid_urls
    try:
        with open(file_path, "r", encoding="utf-8") as file:
            for line in file:
                url = line.strip()
                if any(
                    url.startswith(pattern) for pattern in valid_patterns
                ) and url.endswith(".html"):
                    valid_urls.append(url)
                else:
                    logging.warning(f"[WARNING] URL tidak valid di file: {url}")
        if not valid_urls:
            logging.error("❌⠀ Tidak ada URL valid di file")
            console.print(
                "❌⠀ [red]Tidak ada URL valid di file. Program berhenti.[/red]"
            )
            sys.exit(1)
        return valid_urls
    except Exception as e:
        logging.error(f"❌⠀ Gagal membaca file: {str(e)}")
        console.print(f"❌⠀ [red]Terjadi error saat membaca file: {str(e)}[/red]")
        sys.exit(1)


def select_output_option() -> str:
    """Meminta pengguna memilih opsi pencetakan output"""
    console.print(
        Panel(
            "\n1. Cetak semua data di database \n2. Cetak file individual saja \n3. Cetak keduanya \n4. Simpan data ke database ",
            title="[bold blue]PILIH OPSI FILE OUTPUT[/bold blue]",
            border_style="cyan",
            expand=True,
        )
    )

    while True:
        try:
            choice = console.input("Masukkan nomor opsi (1-4): ").strip()
            if choice in ["1", "2", "3", "4"]:
                return choice
            console.print(
                "⚠️ [yellow]Opsi tidak valid! Harap masukkan nomor 1, 2, 3, atau 4.[/yellow]"
            )
        except Exception as e:
            logging.error(f"❌⠀ Error memilih opsi pencetakan: {str(e)}")
            console.print(
                f"❌⠀ [red]Terjadi error! Harap masukkan nomor 1, 2, 3, atau 4.[/red]"
            )
        time.sleep(1)


def process_single_url(url: str) -> tuple[List, str]:
    """Memproses scraping untuk satu URL, mengembalikan scraped_data dan container_title"""
    target_code = url.split("/")[-1].replace(".html", "")
    logging.info(f"⚙⠀ Memulai proses untuk target_code: {target_code}")

    with BrowserManager() as browser_manager:
        scraper = None
        try:
            for attempt in range(3):
                try:
                    scraper = FileCryptScraper(browser_manager.context.new_page())
                    break
                except Exception as e:
                    logging.warning(
                        f"[WARNING] Gagal membuka halaman baru, mencoba lagi ({attempt + 1}/3): {str(e)}"
                    )
                    time.sleep(2)
            else:
                raise RuntimeError("Gagal membuka halaman baru setelah 3 percobaan")

            for attempt in range(3):
                try:
                    scraper.page.goto(url, wait_until="load", timeout=15000)
                    break
                except Exception as e:
                    logging.warning(
                        f"[WARNING] Gagal memuat URL, mencoba lagi ({attempt + 1}/3): {str(e)}"
                    )
                    time.sleep(2)
            else:
                raise RuntimeError("Gagal memuat URL setelah 3 percobaan")

            browser_manager.close_about_blank_tabs()
            time.sleep(1)

            if not scraper.handle_password():
                raise RuntimeError("Gagal menangani password")
            if not scraper.handle_captcha():
                raise RuntimeError("Gagal menangani CAPTCHA")

            title, total_size = scraper.get_additional_info()
            container_title = (
                title
                if title and title.strip().lower() not in ("n/a", "unknown", "")
                else target_code
            )

            online_count = sum(
                1
                for row in scraper.page.query_selector_all("tr.kwj3")
                if row.query_selector("td.status i")
                and "online" in row.query_selector("td.status i").get_attribute("class")
            )
            total_count = len(scraper.page.query_selector_all("tr.kwj3"))
            print_info(
                {
                    "Judul": title,
                    "Status URL": f"{online_count}/{total_count} Online",
                    "Total Size": total_size,
                    "Target Code": target_code,
                }
            )

            providers = scraper.get_available_providers()
            selected_provider = select_provider(providers)

            logging.info(f"⚙⠀ Memulai proses scraping untuk target_code: {target_code}")
            logging.info(f"⚙⠀ Total item di halaman: {total_count}")
            scraped_data = scraper.scrape_file_info(
                selected_provider=selected_provider,
                all_providers=providers if not selected_provider else None,
            )
            return scraped_data, container_title
        except Exception as e:
            logging.error(f"❌⠀ Gagal memproses URL {url}: {str(e)}")
            return [], target_code
        finally:
            if scraper and scraper.page:
                try:
                    scraper.page.close()
                except Exception as e:
                    logging.warning(f"[WARNING] Gagal menutup halaman: {str(e)}")
            try:
                for page in browser_manager.context.pages:
                    if not page.is_closed():
                        page.close()
            except Exception as e:
                logging.warning(f"[WARNING] Gagal menutup halaman sisa: {str(e)}")


def connect_db(db_name: str) -> Optional[sqlite3.Connection]:
    """Fungsi untuk koneksi ke database"""
    try:
        conn = sqlite3.connect(db_name)
        return conn
    except sqlite3.Error as e:
        console.print(f"❌⠀ [red]Error koneksi database: {e}[/red]")
        logging.error(f"❌⠀ Error koneksi database: {str(e)}")
        return None


def modify_title(title: str) -> str:
    """Fungsi untuk memodifikasi title sesuai pola yang diinginkan"""
    if not title:
        return title

    patterns_to_remove = [
        r"\[MkvDrama\.Org\]",
        r"\[MkvDrama\.me\]",
        r"\.mkv\b",
        r"\.x264\b",
        r"\.1080p\b",
        r"\.720p\b",
        r"\b.E01\b",
    ]
    cleaned_title = title
    for pattern in patterns_to_remove:
        cleaned_title = re.sub(pattern, "", cleaned_title, flags=re.IGNORECASE)

    match = re.match(r"^(.*?S01)(?:E\d{2})?\.?([A-Z]+)?", cleaned_title.strip())
    if match:
        base_title = match.group(1)
        provider = match.group(2) or ""
        return f"{base_title}{'.' + provider if provider else ''}"

    return cleaned_title.strip()


def sanitize_filename(filename: str) -> str:
    """Mengganti karakter tidak valid untuk nama file, tapi pertahankan titik"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, "")
    return filename


def is_sheet_empty(sheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    return sheet.max_row <= 1


def save_to_excel(data: List[tuple], filename: str):
    """Fungsi untuk menyimpan data ke file Excel dengan struktur seperti save_individual_files"""
    try:
        os.makedirs("results", exist_ok=True)
        filename = os.path.join("results", f"{sanitize_filename(filename)}.xlsx")

        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
            for sheet_name in wb.sheetnames[:]:
                sheet = wb[sheet_name]
                if is_sheet_empty(sheet) or sheet_name.lower() == "sheet":
                    wb.remove(sheet)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

        all_sheet_name = "ALL_PROVIDER_RESULTS"
        if all_sheet_name not in wb.sheetnames:
            all_sheet = wb.create_sheet(all_sheet_name, 0)
            all_sheet.append(
                [
                    "No",
                    "Title",
                    "Provider",
                    "Size",
                    "Status",
                    "Download URL",
                    "Bypass URL",
                    "_target_code",
                ]
            )
        else:
            all_sheet = wb[all_sheet_name]

        existing_all_entries = set()
        for row in all_sheet.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2] and row[7]:
                existing_all_entries.add(
                    (
                        str(row[1]).lower().strip(),
                        str(row[2]).lower().strip(),
                        str(row[7]).lower().strip(),
                    )
                )

        new_all_items = 0
        for item in data:
            item_key = (
                str(item[0]).lower().strip(),
                str(item[1]).lower().strip(),
                str(item[6]).lower().strip(),
            )
            if item_key not in existing_all_entries:
                all_sheet.append(
                    [
                        all_sheet.max_row,
                        item[0],
                        item[1],
                        item[2],
                        item[3],
                        item[4],
                        item[5],
                        item[6],
                    ]
                )
                existing_all_entries.add(item_key)
                new_all_items += 1

        for idx, row in enumerate(all_sheet.iter_rows(min_row=2), 1):
            row[0].value = idx

        if is_sheet_empty(all_sheet):
            wb.remove(all_sheet)

        providers = {item[1] for item in data if item[1]}
        for provider in providers:
            provider_items = [item for item in data if item[1] == provider]
            provider_sheet_name = provider[:31]
            existing_provider_entries = set()

            if provider_sheet_name in wb.sheetnames:
                provider_sheet = wb[provider_sheet_name]
                for row in provider_sheet.iter_rows(min_row=2, values_only=True):
                    if row[1] and row[2] and row[7]:
                        existing_provider_entries.add(
                            (
                                str(row[1]).lower().strip(),
                                str(row[2]).lower().strip(),
                                str(row[7]).lower().strip(),
                            )
                        )

            provider_new_items = False
            for item in provider_items:
                item_key = (
                    str(item[0]).lower().strip(),
                    str(item[1]).lower().strip(),
                    str(item[6]).lower().strip(),
                )
                if item_key not in existing_provider_entries:
                    provider_new_items = True
                    break

            if provider_new_items:
                if provider_sheet_name not in wb.sheetnames:
                    provider_sheet = wb.create_sheet(provider_sheet_name)
                    provider_sheet.append(
                        [
                            "No",
                            "Title",
                            "Provider",
                            "Size",
                            "Status",
                            "Download URL",
                            "Bypass URL",
                            "_target_code",
                        ]
                    )
                else:
                    provider_sheet = wb[provider_sheet_name]

                for item in provider_items:
                    item_key = (
                        str(item[0]).lower().strip(),
                        str(item[1]).lower().strip(),
                        str(item[6]).lower().strip(),
                    )
                    if item_key not in existing_provider_entries:
                        provider_sheet.append(
                            [
                                provider_sheet.max_row,
                                item[0],
                                item[1],
                                item[2],
                                item[3],
                                item[4],
                                item[5],
                                item[6],
                            ]
                        )
                        existing_provider_entries.add(item_key)

                for idx, row in enumerate(provider_sheet.iter_rows(min_row=2), 1):
                    row[0].value = idx

                if is_sheet_empty(provider_sheet):
                    wb.remove(provider_sheet)

        has_data = False
        for sheet_name in wb.sheetnames[:]:
            sheet = wb[sheet_name]
            if not is_sheet_empty(sheet):
                has_data = True
            else:
                wb.remove(sheet)

        if not has_data or not wb.sheetnames:
            console.print("⚠️ [yellow]Tidak ada data untuk disimpan.[/yellow]")
            logging.info("[EXCEL] Tidak ada data untuk disimpan ke Excel")
            return

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            column_widths = {
                "A": 5,
                "B": 60,
                "C": 15,
                "D": 10,
                "E": 15,
                "F": 60,
                "G": 60,
                "H": 15,
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(filename)
        console.print(f"✅ [white]Data disimpan ke {filename}[/white]")
        logging.info(f"[EXCEL] Data disimpan ke {filename}")

    except Exception as e:
        console.print(f"❌⠀ [red]Error menyimpan ke Excel: {e}[/red]")
        logging.error(f"❌⠀ Gagal menyimpan ke Excel: {str(e)}")


def save_all_data():
    """Fungsi untuk menyimpan semua data dari database ke RESULT_DATABASE.xlsx"""
    conn = connect_db("results/scraped_data.db")
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT title, provider, size, status, download_url, bypass_url, target_code
                FROM scraped_data
            """
            )
            rows = cursor.fetchall()

            if rows:
                save_to_excel(rows, "RESULT_DATABASE")
            else:
                console.print("⚠️ [yellow]Tidak ada data di database[/yellow]")
                logging.info("📥 Tidak ada data di database")

        except sqlite3.Error as e:
            console.print(f"❌⠀ [red]Error membaca data: {e}[/red]")
            logging.error(f"❌⠀ Error membaca data: {str(e)}")
        finally:
            conn.close()


def display_and_save_by_target_code(target_code: str, modified_title: str):
    """Fungsi untuk menyimpan data dengan target_code tertentu ke Excel"""
    conn = connect_db("results/scraped_data.db")
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT title, provider, size, status, download_url, bypass_url, target_code
                FROM scraped_data
                WHERE target_code = ?
            """,
                (target_code,),
            )
            rows = cursor.fetchall()

            if rows:
                save_to_excel(rows, modified_title)
            else:
                console.print(
                    f"⚠️ [yellow]Tidak ada data dengan target_code: {target_code}[/yellow]"
                )
                logging.info(f"📥 Tidak ada data dengan target_code: {target_code}")

        except sqlite3.Error as e:
            console.print(f"❌⠀ [red]Error membaca data: {e}[/red]")
            logging.error(f"❌⠀ Error membaca data: {str(e)}")
        finally:
            conn.close()


def display_data():
    """Fungsi untuk menampilkan daftar data dalam tabel dan memilih target_code atau semua data"""
    db_name = "results/scraped_data.db"
    conn = connect_db(db_name)

    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT MIN(title), target_code
                FROM scraped_data
                GROUP BY target_code
                """
            )
            rows = cursor.fetchall()

            if not rows:
                console.print("⚠️ [yellow]Tidak ada data di database.[/yellow]")
                logging.info("📥 Tidak ada data di database")
                return

            table = Table(
                title=Text(f"Data dari {db_name}", style="bold blue"),
                show_header=True,
                header_style="bold cyan",
            )
            table.add_column("No", style="cyan", justify="center")
            table.add_column("Title", style="green")
            table.add_column("Target Code", style="magenta")

            selections = []
            for idx, row in enumerate(rows, 1):
                title, target_code = row
                modified_title = modify_title(title)
                table.add_row(str(idx), modified_title, target_code)
                selections.append((modified_title, target_code))

            console.print(table)
            console.print("\n0. Simpan semua data ke RESULT_DATABASE.xlsx")
            console.print("m. Kembali ke menu utama")

            try:
                choice = console.input("\nMasukkan nomor untuk cetak data: ").strip()
                if choice == "0":
                    save_all_data()
                elif choice == "m":
                    logging.debug("Pengguna memilih kembali ke menu utama")
                    main()
                else:
                    choice = int(choice)
                    if 1 <= choice <= len(selections):
                        selected_title, selected_target_code = selections[choice - 1]
                        display_and_save_by_target_code(
                            selected_target_code, selected_title
                        )
                    else:
                        console.print("⚠️ [yellow]Pilihan tidak valid![/yellow]")
                        logging.warning("[INPUT] Pilihan tidak valid")
            except ValueError:
                console.print("⚠️ [yellow]Masukkan angka yang valid![/yellow]")
                logging.warning("[INPUT] Input bukan angka")

        except sqlite3.Error as e:
            console.print(f"❌⠀ [red]Error membaca data: {e}[/red]")
            logging.error(f"❌⠀ Error membaca data: {str(e)}")
        finally:
            conn.close()


def main():
    """Fungsi utama untuk menjalankan scraper"""
    log_file = setup_logging()
    DatabaseHandler._init_db()

    try:
        input_method = select_input_method()
        urls = []

        if input_method == "1":
            urls = [get_valid_url()]
        elif input_method == "2":
            urls = get_urls_from_file()
        elif input_method == "3":
            display_data()
            return

        processed_target_codes = []
        container_titles = {}
        for idx, url in enumerate(urls, 1):
            logging.debug(f"⚙⠀ Memproses URL {idx}/{len(urls)}: {url}")
            console.print(f"🚀 [white]Memproses URL {idx}/{len(urls)}: {url}[/white]")
            scraped_data, container_title = process_single_url(url)

            if scraped_data:
                new_items = DatabaseHandler.save_to_sqlite(scraped_data)
                console.print(
                    f"✅ [white]Berhasil menyimpan {new_items} item baru[/white]"
                )
                if new_items == 0:
                    console.print("⚠️ [yellow]Tidak ada item baru[/yellow]")
                    logging.debug(f"⚙⠀ Tidak ada item baru")
                target_code = url.split("/")[-1].replace(".html", "")
                processed_target_codes.append(target_code)
                container_titles[target_code] = container_title
            else:
                console.print("⚠️ [yellow]Tidak ada data yang di-scrape[/yellow]")
                logging.info(f"⚙⠀ Tidak ada data yang di-scrape")
                console.print(
                    f"⚠️ [yellow]Tidak ada data yang berhasil di-scrape[/yellow]"
                )

            time.sleep(3)

        if urls:
            if processed_target_codes:
                output_option = select_output_option()
                if output_option == "1":
                    all_data = DatabaseHandler.get_all_data()
                    if all_data:
                        DatabaseHandler.export_to_excel(
                            all_data, processed_target_codes[0]
                        )
                        console.print(
                            f"✅ [white]Berhasil menyimpan file database di results\\RESULT_DATABASE.xlsx[/white]"
                        )
                        logging.debug(
                            f"📁⠀Berhasil menyimpan file database di results\\RESULT_DATABASE.xlsx"
                        )
                    else:
                        console.print("⚠️ [yellow]Tidak ada data di database[/yellow]")
                        logging.info("📁⠀Tidak ada data di database")
                elif output_option == "2":
                    for target_code in processed_target_codes:
                        scraped_data = DatabaseHandler.get_data_by_target_code(
                            target_code
                        )
                        if scraped_data:
                            container_title = container_titles.get(
                                target_code, target_code
                            )
                            FileHandler.save_individual_files(
                                scraped_data, target_code, container_title
                            )
                            console.print(
                                f"✅ [white]Berhasil menyimpan file individual di results\\{FileHandler._sanitize_filename(container_title, target_code)}.xlsx[/white]"
                            )
                            logging.debug(
                                f"📁⠀Berhasil menyimpan file individual di results\\{FileHandler._sanitize_filename(container_title, target_code)}.xlsx"
                            )
                elif output_option == "3":
                    all_data = DatabaseHandler.get_all_data()
                    if all_data:
                        DatabaseHandler.export_to_excel(
                            all_data, processed_target_codes[0]
                        )
                        console.print(
                            f"✅ [white]Berhasil menyimpan file database di results\\RESULT_DATABASE.xlsx[/white]"
                        )
                        logging.debug(
                            f"📁⠀Berhasil menyimpan file database di results\\RESULT_DATABASE.xlsx"
                        )
                    for target_code in processed_target_codes:
                        scraped_data = DatabaseHandler.get_data_by_target_code(
                            target_code
                        )
                        if scraped_data:
                            container_title = container_titles.get(
                                target_code, target_code
                            )
                            FileHandler.save_individual_files(
                                scraped_data, target_code, container_title
                            )
                            console.print(
                                f"✅ [white]Berhasil menyimpan file individual di results\\{FileHandler._sanitize_filename(container_title, target_code)}.xlsx[/white]"
                            )
                            logging.debug(
                                f"📁⠀Berhasil menyimpan file individual di results\\{FileHandler._sanitize_filename(container_title, target_code)}.xlsx"
                            )
                elif output_option == "4":
                    console.print(
                        "✅ [white]Data berhasil tersimpan di DATABASE[/white]"
                    )
                    logging.debug("⚙⠀ Data berhasil tersimpan di DATABASE")
            else:
                console.print(
                    "⚠️ [yellow]Tidak ada data yang berhasil di-scrape dari URL manapun[/yellow]"
                )
                logging.debug(
                    "⚙⠀ Tidak ada data yang berhasil di-scrape dari URL manapun"
                )
                console.print(
                    "⚠️ [yellow]Tidak ada data yang berhasil di-scrape dari URL manapun.[/yellow]"
                )

        console.print(
            "🎉 [white]Scraping selesai. Output telah diproses. Program selesai.[/white]"
        )
        logging.debug("⚙⠀ Scraping selesai. Output telah diproses. Program selesai.")

    except Exception as e:
        console.print(f"❌⠀ [red]Terjadi kesalahan: {str(e)}[/red]")
        logging.error(f"❌⠀ Terjadi kesalahan: {str(e)}")
        sys.exit(1)


def print_info(info: Dict[str, str]):
    """Menampilkan informasi URL dengan rich"""
    content = "\n".join(
        f"[cyan]{key:<15}[/cyan] : [white]{value}[/white]"
        for key, value in info.items()
    )
    console.print(
        Panel(
            content,
            title="[bold blue]INFORMASI URL[/bold blue]",
            border_style="cyan",
            expand=True,
        )
    )


def select_provider(providers: List[str]) -> Optional[str]:
    if not providers:
        console.print("⚠️ [yellow]Tidak ada provider yang tersedia.[/yellow]")
        return None

    content = "\n".join(
        f"[white]{i}. {provider}[/white]" for i, provider in enumerate(providers, 1)
    )
    console.print(
        Panel(
            content,
            title="[bold blue]PROVIDER YANG TERSEDIA[/bold blue]",
            border_style="cyan",
            expand=True,
        )
    )

    try:
        provider_input = console.input(
            "Masukkan nomor provider (0 untuk semua): "
        ).strip()
        if provider_input == "0":
            return None
        elif provider_input.isdigit():
            provider_num = int(provider_input)
            if 1 <= provider_num <= len(providers):
                return providers[provider_num - 1]
        console.print(
            "⚠️ [yellow]Nomor tidak valid! Akan mengambil semua provider.[/yellow]"
        )
        time.sleep(2)
        return None
    except Exception as e:
        console.print(f"❌⠀ [red]Error memilih provider: {str(e)}[/red]")
        logging.error(f"❌⠀ Error memilih provider: {str(e)}")
        console.print(
            "⚠️ [yellow]Terjadi error! Akan mengambil semua provider.[/yellow]"
        )
        time.sleep(2)
        return None


if __name__ == "__main__":
    try:
        for ext in DEFAULT_CONFIG.extensions.paths:
            if not os.path.exists(os.path.join(ext, "manifest.json")):
                raise FileNotFoundError(f"❌⠀ Ekstensi tidak valid di: {ext}")

        console.print(
            Panel(
                Text("FILECRYPT SCRAPER", style="bold blue", justify="center"),
                title="",
                border_style="cyan",
                expand=True,
            )
        )

        main()
        console.print(
            Panel(
                Text("FILECRYPT SELESAI", style="bold green", justify="center"),
                title="",
                border_style="cyan",
                expand=True,
            )
        )
    except Exception as e:
        console.print(f"❌⠀ [red]Error: {str(e)}[/red]")
        logging.error(f"❌⠀ Error: {str(e)}")
        sys.exit(1)
